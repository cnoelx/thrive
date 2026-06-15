from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = 'Arial'
INK = '0C1410'
HEAD_FILL = PatternFill('solid', fgColor='14532D')
HEAD_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=11)
CELL_FONT = Font(name=FONT, color=INK, size=10)
WHY_FONT = Font(name=FONT, color='6B7280', size=10, italic=True)
TITLE_FONT = Font(name=FONT, bold=True, color=INK, size=15)
thin = Side(style='thin', color='E6E8EC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')
top = Alignment(vertical='top')

# (category, name, why, sets, [L1..])
EX = [
    ('Move', 'Bodyweight Squat', 'Stand, sit, lower all day', 2,['15','22','30','40','50','tempo (4s lower) ×8','paused (3s bottom) ×8','split squat ×8/leg','Bulgarian split ×6/leg','pistol progression ×3/leg']),
    ('Move', 'Reverse Lunge', 'Stairs, single-leg strength (per leg)', 2,['10 (5/leg)','16 (8/leg)','24 (12/leg)','32 (16/leg)','40 (20/leg)','walking ×12/leg','deficit ×8/leg','Bulgarian split ×8/leg','jumping lunge ×6/leg','pistol progression ×3/leg']),
    ('Move', 'Sit-to-Stand from Floor', 'Get off the ground unaided (one-time check)', None, ['Hands OK','Fewer hands','One hand','One fingertip','No hands']),
    ('Move', 'Single-leg Balance', "Don't fall", 2, ['10s/leg','20s/leg','30s/leg','45s/leg','60s/leg','eyes-closed 20s','eyes-closed 40s','unstable surface 30s']),
    ('Move', 'Plank', 'Stable trunk, protects back', 2,['20s','35s','50s','70s','90s','120s','RKC tension 30s','single-arm 20s','single-arm+leg 15s']),
    ('Move', 'Side Plank', 'Side core — resists twisting', 2,['knee 10s/side','knee 20s/side','full 30s/side','full 45s/side','full 60s/side','full 75s/side','star side plank 20s','+reach-through ×10','single-leg side plank 20s']),
    ('Move', 'Glute Bridge', 'Glutes & hamstrings (backside)', 2,['10','16','24','34','single-leg ×8/leg','single-leg ×12/leg','feet-elevated SL ×10/leg','marching hold 30s','hip thrust (load) ×12','single-leg hip thrust ×8/leg']),
    ('Move', 'Prone Leg Raise', 'Glutes & hamstrings', 2, ['6/leg','10/leg','15/leg','20/leg','20/leg + 2s hold','both legs + hold ×12','swimmer flutter 30s','+light ankle load ×12','superman rock 30s']),
    ('Push', 'Push-ups', 'Push, catch a fall', 2,['wall ×8','incline ×8 (hands on a counter)','incline ×8 (hands on a low step)','negatives ×5 (lower down slowly)','5 full push-ups','8 full push-ups','feet-elevated ×8','diamond ×8','archer ×5/side','one-arm progression ×3/side']),
    ('Pull', 'Inverted Row', 'Horizontal pull (rings/TRX, or bar set low). Angle = difficulty.', 2,['upright, 5 reps','upright, 8 reps','leaned back, 10 reps','leaned back, 12 reps','near-horizontal, 12 reps','feet-elevated ×10','tuck front-lever raises ×5','tuck front-lever hold 8s','wide front-lever tuck 10s','one-arm row progression ×5/side']),
    ('Pull', 'Pulling', 'Pull your bodyweight up — hang, then pull (needs bar or rings)', 2,['active hang 15s','scapular pulls ×5','negatives ×3 (jump to the top, lower for 5s)','5 with a band','1 pull-up','3 pull-ups','6 pull-ups','chest-to-bar ×5','archer ×3/side','one-arm progression / weighted']),
    ('Cardio', 'Walk / Run (walk-first)', 'Cover distance without gassing out', None, ['brisk walk 10min','brisk walk 20min','brisk walk 30min','walk-jog 20min','easy jog 20min+','jog 30min','jog 5km','intervals (speed)','5km steady pace','10km / sustained']),
    ('Mobility', 'Deep Squat Hold', 'Rest, garden, floor play (one-time check)', None, ['support 15s','support 30s','free 30s','free 60s','free 90s']),
    ('Mobility', 'Overhead Reach', 'Shelves, posture — wall test (one-time check)', None, ['ribs stick out a lot','ribs stick out a little','ribs stay down','ribs down + slight backbend','full, pain-free']),
    ('Mobility', 'Ankle Mobility', 'Squat deep, descend safely — knee-to-wall (one-time check)', None, ['heel lifts','slight lift','knee over toes','knee well past','no heel lift']),
]

SCHEDULE = [
    ('Monday', 'Full Body', [('Bodyweight Squat', 2,60, ''), ('Push-ups', 2,60, ''), ('Inverted Row', 2,60, ''), ('Plank', 2,45, '')]),
    ('Tuesday', 'Cardio & Mobility', [('Walk / Run', None, None, '20–30 min, conversational pace'), ('Deep Squat Hold', 2, 30, 'check'), ('Ankle Mobility', 1, None, 'knee-to-wall, both sides'), ('Overhead Reach', 1, None, 'wall test')]),
    ('Wednesday', 'Legs & Core', [('Reverse Lunge', 2,60, ''), ('Glute Bridge', 2,45, ''), ('Prone Leg Raise', 2, 45, ''), ('Side Plank', 2,45, '')]),
    ('Thursday', 'Upper Body', [('Pulling', 2,90, ''), ('Push-ups', 2, 60, 'lighter'), ('Single-leg Balance', 2, 30, '')]),
    ('Friday', 'Strength & Core', [('Bodyweight Squat', 2,60, ''), ('Inverted Row', 2,60, ''), ('Glute Bridge', 2,45, ''), ('Plank', 2,45, '')]),
    ('Saturday', 'Cardio', [('Walk / Run', None, None, '20 min — walk or walk-jog intervals')]),
    ('Sunday', 'Rest', []),
]

NOTES = [
    'Levels are tracked PER CATEGORY (L1–L10). A category completes a level when every exercise in it is claimed at that level.',
    'Max level is 10. Mobility is a checkpoint category and caps at Level 5. Some exercises cap earlier (Balance L8; Plank, Side Plank, Prone Leg Raise L9).',
    'Claims are self-reported — you tick a goal off once you can do it with good form.',
    'Runway: a category can only progress to (overall level + 2) — no sprinting more than two levels ahead of the rest.',
    'Workouts: reps/time come from your current level and auto-evolve as you level up. Sets are capped at 2 per exercise.',
    'When chasing the next level, set 1 is a fresh goal attempt; the remaining set trains the completed level (work set).',
    'Checkpoints (Sit-to-Stand, Overhead Reach, Ankle Mobility) are one-time form checks, not rep counts.',
    'Pull (Inverted Row + Pulling) needs a bar or rings. Without equipment, Superman substitutes and Pull is left out of progress.',
    'Source of truth: data/benchmarks.ts (ladders) and data/schedule.ts (weekly plan). Generated 2026-06-12.',
]

wb = Workbook()

# ---- Exercises ----
ws = wb.active
ws.title = 'Exercises'
ws['A1'] = 'Thrive — Exercise Ladders'
ws['A1'].font = TITLE_FONT
ws.append([])
headers = ['Category', 'Exercise', 'Why it matters', 'Sets'] + [f'L{i}' for i in range(1, 11)]
ws.append(headers)
hrow = ws.max_row
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=hrow, column=c)
    cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = BORDER
for cat, name, why, sets, targets in EX:
    row = [cat, name, why, ('—' if sets is None else sets)] + targets + [''] * (10 - len(targets))
    ws.append(row)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER; cell.alignment = wrap
        cell.font = WHY_FONT if c == 3 else CELL_FONT
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='center', vertical='top')
ws.freeze_panes = 'A4'
widths = [11, 22, 30, 6] + [16] * 10
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

# ---- Weekly Schedule ----
ws2 = wb.create_sheet('Weekly Schedule')
ws2['A1'] = 'Thrive — Weekly Schedule'
ws2['A1'].font = TITLE_FONT
ws2.append([])
h2 = ['Day', 'Focus', 'Exercise', 'Sets', 'Rest (s)', 'Note']
ws2.append(h2)
hrow2 = ws2.max_row
for c in range(1, len(h2) + 1):
    cell = ws2.cell(row=hrow2, column=c)
    cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = BORDER
for day, focus, items in SCHEDULE:
    if not items:
        ws2.append([day, focus, '—', '', '', 'Rest day'])
        for c in range(1, 7):
            ws2.cell(row=ws2.max_row, column=c).font = CELL_FONT; ws2.cell(row=ws2.max_row, column=c).border = BORDER
        continue
    for idx, (ex, sets, rest, note) in enumerate(items):
        ws2.append([day if idx == 0 else '', focus if idx == 0 else '', ex, ('—' if sets is None else sets), ('—' if rest is None else rest), note])
        r = ws2.max_row
        for c in range(1, 7):
            cell = ws2.cell(row=r, column=c); cell.font = CELL_FONT; cell.border = BORDER; cell.alignment = top
        ws2.cell(row=r, column=4).alignment = Alignment(horizontal='center', vertical='top')
        ws2.cell(row=r, column=5).alignment = Alignment(horizontal='center', vertical='top')
    ws2.cell(row=ws2.max_row - len(items) + 1, column=1).font = Font(name=FONT, bold=True, color=INK, size=10)
ws2.freeze_panes = 'A4'
for col, w in zip('ABCDEF', [12, 16, 26, 6, 9, 30]):
    ws2.column_dimensions[col].width = w

# ---- Notes ----
ws3 = wb.create_sheet('Notes')
ws3['A1'] = 'How the program works'
ws3['A1'].font = TITLE_FONT
ws3.append([])
for n in NOTES:
    ws3.append(['•', n])
    r = ws3.max_row
    ws3.cell(row=r, column=1).font = CELL_FONT; ws3.cell(row=r, column=1).alignment = top
    ws3.cell(row=r, column=2).font = CELL_FONT; ws3.cell(row=r, column=2).alignment = wrap
ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 92

import os
out = os.path.join(os.path.expanduser('~'), 'Downloads', 'Functional_Fitness_Program_rebuilt_2.xlsx')
wb.save(out)
print('SAVED', out)
